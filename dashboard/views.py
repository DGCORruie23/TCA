from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.db.models import Max
# from usuarios.models import usuarioL
from usuarios.models import UsuarioP
from usuarios.models import Registro, Acciones, Notificacion,Area, Rubro
from datetime import datetime, timedelta, timezone
from .forms import RegistroConAccionesYPruebasForm, MensajeForm, AccionesForm, RegistroConAccionesFORM, CargarArchivoForm
from django.forms import inlineformset_factory
import openpyxl as opxl

import locale
from datetime import date
from datetime import datetime
import re
from django.contrib import messages

from django.db import transaction
from dateutil.parser import parse

from twilio.rest import Client

import os
from dotenv import load_dotenv

load_dotenv()

@login_required
def dashboard(request):
    if request.method == 'GET':
        userDataI = UsuarioP.objects.filter(user__username=request.user)
        registrosConFechas = []
        nombres_areas = [area.nickname for area in Area.objects.all()]
        areas_n = nombres_areas if len(nombres_areas) > 1 else None
    

        # print(areas_n)
        formCargar1 = CargarArchivoForm()
        if userDataI[0].tipo == "1":
            registros = Registro.objects.all().order_by('fecha_termino')
        else:
            registros_area = Registro.objects.filter(area=userDataI[0].OR).order_by('fecha_termino')
            registros_acciones_area2 = Registro.objects.filter(
                accionR__area2=userDataI[0].OR
            ).order_by('fecha_termino')

            registros = registros_area | registros_acciones_area2
            registros = registros.distinct().order_by('fecha_termino')

        registros_en_proceso = registros.filter(estado="1")
        registros_atendidos = registros.filter(estado="2")

        registros_ordenados = list(registros_en_proceso) + list(registros_atendidos)

        for registro in registros_ordenados:
            fecha_inicio_str = registro.fecha_inicio.strftime('%d-%m-%Y')
            fecha_inicio = fecha_inicio_str.split('-')

            fecha_termino_str = registro.fecha_termino.strftime('%d-%m-%Y')
            fecha_termino = fecha_termino_str.split('-')

            fecha_inicio_dt = datetime.strptime(fecha_inicio_str, '%d-%m-%Y')
            fecha_termino_dt = datetime.strptime(fecha_termino_str, '%d-%m-%Y')
            diferencia = datetime.now() - fecha_termino_dt

            fecha_finalizacion = registro.fecha_finalizacion

            areas = registro.area.all()
            areas_str = ', '.join(area.nickname for area in areas)
            areas_name = ', '.join(area.name for area in areas)

            dias = diferencia.days
            porcentaje = registro.porcentaje_avance
            clave_acuerdo_partes = registro.claveAcuerdo.split('/')


            registrosConFechas.append({
                'registro': registro,
                'fecha_inicio': fecha_inicio,
                'fecha_termino': fecha_termino,
                'diferencia': dias,
                'areas_str': areas_str,
                'areas_name': areas_name,
                'fecha_finalizacion': fecha_finalizacion,
                'porcentaje': porcentaje,
                'clave_acuerdo_partes': clave_acuerdo_partes
            })

        now = datetime.now()
        nuevos_registros = registros.filter(fecha_creacion__gte=now - timedelta(days=365))


        for registro in nuevos_registros:
            Notificacion.objects.get_or_create(user=request.user, registro=registro)

        notificaciones = Notificacion.objects.filter(user=request.user, leido=False)

        context = {
            'registrosConFechas': registrosConFechas,
            'dataU': userDataI,
            'notificaciones': notificaciones,
            'formCargar1': formCargar1,
            'areas_n': areas_n
        }

        return render(request, "dashboard/dashboard.html", context)

@login_required
def marcar_notificacion_leida(request, notificacion_id):
    notificacion = get_object_or_404(Notificacion, id=notificacion_id, user=request.user)
    notificacion.leido = True
    notificacion.fecha_leido = datetime.now()
    notificacion.save()

    registro_id = notificacion.registro.idRegistro
    return redirect('detalles', registro_id=registro_id)



@login_required
def detalles(request, registro_id):
    registro = Registro.objects.get(pk=registro_id)
    mensajes = registro.mensajes.all().order_by('fecha_envio')
    userDataI = UsuarioP.objects.filter(user__username=request.user)

    for mensaje in mensajes:
        if mensaje.archivo:
            mensaje.archivo_nombre = mensaje.archivo.name.split('/')[-1]

    if request.method == 'POST':
        mensaje_form = MensajeForm(request.POST, request.FILES)
        if mensaje_form.is_valid():
            nuevo_mensaje = mensaje_form.save(commit=False)
            nuevo_mensaje.registro = registro
            nuevo_mensaje.usuario = request.user
            nuevo_mensaje.save()
            return redirect('detalles', registro_id=registro_id)
    else:
        mensaje_form = MensajeForm()

    context = {
        'registro': registro,
        'mensajes': mensajes,
        'mensaje_form': mensaje_form,
        'dataU': userDataI,
    }

    return render(request, 'dashboard/detalles.html', context)





@login_required
def crear_registro(request):
    if request.method == 'POST':
        registro_form = RegistroConAccionesYPruebasForm(request.POST)
        if registro_form.is_valid():
            registro = registro_form.save()

            areas2 = registro_form.cleaned_data['accion1_area2']
            area = registro_form.cleaned_data['area']

            accion = Acciones.objects.create(
                descripcion=registro_form.cleaned_data['accion1_descripcion']
            )

            users = UsuarioP.objects.filter(OR__in=areas2)
            users2 = UsuarioP.objects.filter(OR__in=area)
            print("users")
            for user in users:
                
                print(user.Ntelefono)
            print("users2")
            for user in users2:
                print("users2")
                print(user.Ntelefono)

            accion.area2.set(areas2)
            accion.save()
            registro.accionR.add(accion)

            account_sid = os.getenv('TWILIO_ACCOUNT_SID')
            auth_token = os.getenv('TWILIO_AUTH_TOKEN')

            client = Client(account_sid, auth_token)
            message = client.messages.create(
                from_='whatsapp:+14155238886',
                body='DGCOR te ha añadido al TCA, visita tca.dgcor.com',
                to=f'whatsapp:+5215572247394'
            )

            print(message.sid)

            return redirect('dashboard')
    else:
        registro_form = RegistroConAccionesYPruebasForm()

    return render(request, 'dashboard/crear_registro.html', {
        'registro_form': registro_form,
    })


@login_required
def editar_registro(request, id):
    registro = get_object_or_404(Registro, idRegistro=id)
    accion = registro.accionR.first()
    userDataI = UsuarioP.objects.filter(user__username=request.user)

    if request.method == 'POST':
        registro_form = RegistroConAccionesFORM(request.POST, instance=registro)
        accion_form = AccionesForm(request.POST, instance=accion)

        if registro_form.is_valid() and accion_form.is_valid():
            if registro.estado == 1:
                registro.fecha_finalizacion = "1970-01-01"
            else:
                registro.fecha_finalizacion = datetime.now().strftime("%Y-%m-%d")
            registro = registro_form.save()
            accion = accion_form.save(commit=False)
            accion.registro = registro
            accion.save()

            messages.success(request, "Guardado correctamente.")
            return render(request, 'dashboard/editar_registro.html', {
                'registro_form': registro_form,
                'accion_form': accion_form,
                'dataU': userDataI,

            })

        else:

            for field, errors in registro_form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
            for field, errors in accion_form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        registro_form = RegistroConAccionesFORM(instance=registro)
        accion_form = AccionesForm(instance=accion)

    return render(request, 'dashboard/editar_registro.html', {
        'registro_form': registro_form,
        'accion_form': accion_form,
        'dataU': userDataI,
    })


@login_required
def eliminar_registro(request, idRegistro):
    registro = get_object_or_404(Registro, idRegistro=idRegistro)
    accion = registro.accionR.first()
    userDataI = UsuarioP.objects.filter(user__username=request.user)

    if userDataI[0].tipo == "1" and request.method == 'POST':
        registro.delete()
        accion.delete()
        return redirect('dashboard')

    return render(request, 'dashboard/eliminar_registro.html', {'registro': registro})




locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

def convert_spanish_date(date_str):
    try:
        date_str = date_str.lower()
        return datetime.strptime(date_str, '%d de %B de %Y').date()
    except ValueError:
        print(f"Error al convertir fecha: {date_str}")
        return None



@login_required
@csrf_exempt
def cargaMasiva(request):
    if request.method == "POST":
        form = CargarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()

            print(f"Archivo recibido: {nombreA}")
            errores = 0
            errArea = 0
            if extensionA in ["xlsx", "xlsm", "xlsb", "xltx", "xltm", "xls"]:
                dataWB = opxl.load_workbook(excel_file, data_only=True)
                data = dataWB.worksheets[0]

                total_rows = data.max_row - 1
                chunk_size = 20

                for start_row in range(1, total_rows + 1, chunk_size):
                    end_row = min(start_row + chunk_size, total_rows + 1)
                    
                    with transaction.atomic():
                        for i in range(start_row, end_row):
                            if data.cell(i + 1, 3).value is None:
                                break
                            fecha_inicio = data.cell(i + 1, 1).value
                            fecha_inicio = convert_spanish_date(fecha_inicio) if isinstance(fecha_inicio, str) else fecha_inicio

                            area_celda = data.cell(i + 1, 2).value
                            clave_acuerdo = data.cell(i + 1, 3).value
                            rubro = data.cell(i + 1, 4).value
                            descripcion = data.cell(i + 1, 5).value
                            areas_responsables = data.cell(i + 1, 6).value
                            fecha_termino = data.cell(i + 1, 7).value
                            estado = data.cell(i + 1, 8).value
                            if estado in ["ATENDIDA","Atendida","Atendido", "atendido", "ATENDIDO", "Completado", "COMPLETADO", "Terminado", "TERMINADO", "Cumplido", "CUMPLIDO", "cumplido"]:
                                estado = 2
                            else:
                                estado = 1
                            fecha_finalizacion = data.cell(i + 1, 9).value
                            if fecha_finalizacion is None:
                                fecha_finalizacion = datetime(1970, 1, 1).date()
                            if not isinstance(fecha_termino, date):
                                fecha_termino = fecha_inicio

                            try:
                                rubro_obj = Rubro.objects.get(tipo=rubro)
                            except Rubro.DoesNotExist:
                                errores += 1

                            try:
                                area_obj = Area.objects.get(nickname=area_celda)
                            except Area.DoesNotExist:
                                pass
                                print(f"Área '{area_celda}' no encontrada")

                            areas_responsables_list = areas_responsables.split('-')
                            areas_objs = []
                            for area in areas_responsables_list:
                                try:
                                    if area == "OR":
                                        area_responsable_obj = Area.objects.get(nickname=area_celda)
                                    else:
                                        area_responsable_obj = Area.objects.get(nickname=area.strip())
                                    print(f"Área responsable seteada '{area_responsable_obj}'")
                                except Area.DoesNotExist:
                                    errArea += 1
                                    print(f"Área responsable '{area_responsable_obj}' no encontrada")
                                    pass
                                areas_objs.append(area_responsable_obj)

                            ultimo_clave_acuerdo = Registro.objects.filter(area=area_obj).aggregate(Max('claveAcuerdo'))['claveAcuerdo__max']
                            if ultimo_clave_acuerdo:
                                ultimo_indice = int(ultimo_clave_acuerdo.split('/')[0])
                            else:
                                ultimo_indice = 0
                            nuevo_indice = ultimo_indice + 1
                            clave_acuerdo = f"{nuevo_indice:02}/{area_celda.split(' ')[1]}/{fecha_inicio.strftime('%m/%Y')}"

                            registro, created = Registro.objects.update_or_create(
                                claveAcuerdo=clave_acuerdo,
                                defaults={
                                    "fecha_inicio": fecha_inicio,
                                    "fecha_termino": fecha_termino,
                                    "estado": estado,
                                    "fecha_finalizacion": fecha_finalizacion,
                                }
                            )
                            registro.rubro.set([rubro_obj])
                            registro.area.set([area_obj])

                            accion, created = Acciones.objects.update_or_create(
                                descripcion=descripcion,
                            )
                            accion.idRegistro.add(registro)
                            accion.area2.set(areas_objs)

                            print(f"Registro {i}: {clave_acuerdo} procesado correctamente")
                            print('Errores de area', errArea)

            return redirect('dashboard')
        else:
            print("Formulario no válido")
    else:
        form = CargarArchivoForm()
    return render(request, "dashboard/dashboard.html", {"form": form})
